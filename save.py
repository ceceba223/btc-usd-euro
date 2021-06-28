from openpyxl import Workbook, load_workbook
from datetime import date, datetime

def save(BTC_USD, BTC_EUR, USD_EUR):
    workbook = load_workbook('data.xlsx')
    sheet = workbook.active

    date = datetime.now()

    row = (date, BTC_USD, BTC_EUR, USD_EUR)
    sheet.append(row)

    workbook.save(filename="data.xlsx")

def setup():
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "date"
    sheet["B1"] = "1 BTC in USD"
    sheet["C1"] = "1 BTC in USD"
    sheet["D1"] = "1 USD in EUR"

    workbook.save(filename="data.xlsx")
